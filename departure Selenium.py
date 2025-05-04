from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
import os
import re
from bs4 import BeautifulSoup
import requests
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from datetime import datetime

# 创建 WebDriver 实例并自动管理 ChromeDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

try:
    # 打开目标网页
    driver.get('https://www.flightradar24.com/airport/szx/departures')
    
    # 等待页面加载完成（可以根据实际情况调整等待时间）
    time.sleep(3)
    
    # 模拟点击“Accept Cookies”按钮
    try:
        accept_cookies_button = driver.find_element(By.ID, 'onetrust-accept-btn-handler')
        accept_cookies_button.click()
        print("Cookies 接受按钮已点击")
    except Exception as e:
        print(f"无法找到或点击 Cookies 接受按钮: {e}")
    
    # 等待页面加载完成（可以根据实际情况调整等待时间）
    time.sleep(2)
    
    # 模拟点击“Load later flights”按钮并重复N次
    for _ in range(4):
        try:
            # 查找按钮元素
            button = driver.find_element(By.CSS_SELECTOR, 'button[data-testid="airport-arrival-departure__load-later-flights"]')
            
            # 点击按钮
            button.click()
            
            # 等待页面加载完成（可以根据实际情况调整等待时间）
            time.sleep(2)
        except Exception as e:
            print(f"无法找到或点击按钮: {e}")
            break
    
    # 获取网页源代码
    page_source = driver.page_source
    
    # 将网页源代码保存到文件
    with open('flightradar24_szx_dearture.html', 'w', encoding='utf-8') as file:
        file.write(page_source)
        
    print("网页源代码已成功保存")


finally:
    # 关闭浏览器
    driver.quit()
    
# 关键字列表
keywords = ['74', '75', '76', '77', '78', '33', '35', '36', '38']

def log(message):
    print(f"[LOG] {message}")

def download_image(image_url, folder_path, filename):
    image_path = os.path.join(folder_path, filename)
    
    # 检查文件是否已存在
    if os.path.exists(image_path):
        log(f"Image already exists: {filename}")
        return image_path
    
    try:
        response = requests.get(image_url)
        response.raise_for_status()
        with open(image_path, 'wb') as file:
            file.write(response.content)
        log(f"Downloaded image: {filename}")
        return image_path
    except Exception as e:
        log(f"Error downloading image {image_url}: {e}")
        return None

def extract_info_from_file(file_path, output_data, images_folder):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            soup = BeautifulSoup(file, 'lxml')
            
            # 查找目标li元素
            flight_list_items = soup.find_all('li', class_='airport__flight-list-item')
            
            for item in flight_list_items:
                # 提取航班时间
                time_formatter_element = item.find('div', {'data-testid': 'base-day-period-formatter'})
                if not time_formatter_element:
                    continue
                time_text = time_formatter_element.get_text(strip=True)
                
                # 提取航空型号
                aircraft_model_element = item.find('span', class_='inline-flex h-4 items-center rounded px-1 text-2xs font-semibold bg-blue-200 text-blue-600')
                if not aircraft_model_element:
                    continue
                aircraft_model_text = aircraft_model_element.get_text(strip=True).split()[0]  # 提取型号部分
                
                # 提取日期
                date_element = item.find_previous('h3', class_='inline-flex items-center text-sm uppercase')
                if not date_element:
                    continue
                date_text = date_element.get_text(strip=True)
                
                # 提取航班号
                flight_number_element = item.find('span', class_='truncate text-sm text-gray-900')
                if not flight_number_element:
                    continue
                flight_number_text = flight_number_element.get_text(strip=True).split()[0]  # 提取航班号部分
                
                # 提取图片URL
                logo_div = item.find('div', {'aria-label': 'logo', 'role': 'img'})
                if not logo_div:
                    continue
                style = logo_div.get('style')
                if not style:
                    continue
                match = re.search(r'url\("([^"]+)"\)', style)
                if not match:
                    continue
                image_url = match.group(1)
                
                if any(keyword in aircraft_model_text for keyword in keywords) and "738" not in aircraft_model_text and "733" not in aircraft_model_text and "B38M" not in aircraft_model_text:
                    log(f"Found matching entry in file: {file_path}")
                    log(f"Aircraft Model: {aircraft_model_text}")
                    log(f"Scheduled Departure: {time_text}")
                    log(f"Date: {date_text}")
                    log(f"Flight Number: {flight_number_text}")
                    log(f"Image URL: {image_url}")
                    
                    # 下载图片
                    image_filename = os.path.basename(image_url)
                    image_path = download_image(image_url, images_folder, image_filename)
                    
                    # 添加数据到输出列表
                    output_data.append({
                        'Aircraft Model': aircraft_model_text,
                        'Scheduled Departure': time_text,
                        'Date': date_text,
                        'Flight Number': flight_number_text,
                        'Image Path': image_path
                    })
    except Exception as e:
        log(f"Error processing file {file_path}: {e}")

def save_to_excel(output_data, excel_file_path, images_folder):
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Information"
    
    headers = ['Aircraft Model', 'Scheduled Departure', 'Date', 'Flight Number', 'Image']
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    row_offset = 1  # 表头占用一行
    
    for data in output_data:
        row = [
            data['Aircraft Model'],
            data['Scheduled Departure'],
            data['Date'],
            data['Flight Number'],
            ""
        ]
        ws.append(row)
        
        # 插入图片
        if data['Image Path']:
            img = ExcelImage(data['Image Path'])
            img.width = 50
            img.height = 50
            cell = ws.cell(row=ws.max_row, column=headers.index('Image') + 1)
            ws.add_image(img, f"{cell.coordinate}")
            # 调整行高以适应图片
            ws.row_dimensions[cell.row].height = 50
        
        row_offset += 1
    
    # 自动调整列宽
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    
    wb.save(excel_file_path)

def ensure_unique_excel_file_path(base_name):
    base_path = os.path.splitext(base_name)[0]
    extension = os.path.splitext(base_name)[1]
    counter = 1
    new_file_path = base_name
    
    while os.path.exists(new_file_path):
        new_file_path = f"{base_path}_{counter}{extension}"
        counter += 1
    
    return new_file_path

def rename_existing_file(output_file):
    if os.path.exists(output_file):
        new_output_file = f"output_{timestamp}.xlsx"
        os.rename(output_file, new_output_file)
        log(f"Renamed existing file to: {new_output_file}")

def main():
    current_directory = os.getcwd()
    html_files = [f for f in os.listdir(current_directory) if f.endswith('.html')]
    
    if not html_files:
        log("No HTML files found in the current directory.")
        return
    
    log(f"Found {len(html_files)} HTML files to process.")
    
    output_data = []
    images_folder = os.path.join(current_directory, 'images')
    os.makedirs(images_folder, exist_ok=True)
    
    for html_file in html_files:
        file_path = os.path.join(current_directory, html_file)
        log(f"Processing file: {file_path}")
        extract_info_from_file(file_path, output_data, images_folder)
    
    # 确保Excel文件路径唯一
    excel_file_path = "extracted_flight_info.xlsx"
    unique_excel_file_path = ensure_unique_excel_file_path(excel_file_path)
    
    # 重命名现有文件
    rename_existing_file(unique_excel_file_path)
    
    # 将数据保存到Excel文件
    save_to_excel(output_data, unique_excel_file_path, images_folder)
    
    log(f"Extraction complete.")
    log(f"Results saved to {unique_excel_file_path}.")
    log(f"Images saved to {images_folder}.")

if __name__ == "__main__":
    main()



